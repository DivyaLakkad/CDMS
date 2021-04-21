import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook



wb_data = load_workbook(r'C:\Users\divyal\Documents\GitHub\Front_End_NOVA_TS\Meta_Data.xlsx')
ws_data = wb_data['data_2021']
unit = "Civil"
for row in ws_data.iter_cols(1):
    for cell in row:
        if cell.value == unit:
            print(ws_data.cell(row=cell.row, column=3).value)